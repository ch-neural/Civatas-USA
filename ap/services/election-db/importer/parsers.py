"""CEC (中央選舉委員會) election data parsers.

Handles three main formats from the CEC election database:
1. XLS polling-station detail (投開票所明細) — Big5 encoded
2. XLSX candidate vote summary (候選人得票數一覽表) — UTF-8
3. CSV/TXT aggregated summaries — Big5 encoded

All parsers produce a standardised list of ParsedRow dicts.
"""
from __future__ import annotations

import csv
import io
import logging
import os
import re
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)


# ── Shared helpers ──────────────────────────────────────────

def _clean_number(val) -> int:
    """Parse a number that may contain commas, quotes, or be a float."""
    if val is None:
        return 0
    s = str(val).strip().replace(",", "").replace('"', "").replace("'", "")
    if not s or s == "None":
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def _clean_float(val) -> float | None:
    """Parse a float that may contain commas."""
    if val is None:
        return None
    s = str(val).strip().replace(",", "").replace('"', "")
    if not s or s == "None":
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _strip_ws(s: str) -> str:
    """Strip fullwidth & halfwidth whitespace."""
    return s.strip().strip("\u3000").strip()


def _detect_encoding(filepath: str) -> str:
    """Try common Taiwan encodings."""
    for enc in ("utf-8-sig", "utf-8", "big5", "cp950"):
        try:
            with open(filepath, encoding=enc) as f:
                f.read(2048)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return "utf-8"


# ── Election metadata extraction from filenames ─────────────

# e.g. "111年臺中市縣市長選舉" or "第16任總統副總統"
_YEAR_RE = re.compile(r"(\d{2,3})年")
_AD_YEAR_RE = re.compile(r"(19|20)\d{2}")
_COUNTY_RE = re.compile(r"(臺北市|新北市|桃園市|臺中市|臺南市|高雄市|基隆市|新竹市|嘉義市"
                         r"|新竹縣|苗栗縣|彰化縣|南投縣|雲林縣|嘉義縣|屏東縣|宜蘭縣"
                         r"|花蓮縣|臺東縣|澎湖縣|金門縣|連江縣)")
_CODE_RE = re.compile(r"(\d{5})")  # e.g. 66000


@dataclass
class ElectionMeta:
    """Metadata extracted from filename or sheet title."""
    name: str = ""
    election_type: str = "mayor"       # matches election_type enum
    roc_year: int = 0
    ad_year: int = 0
    scope: str = ""                    # county name or '全國'
    region_code: str = ""
    source_file: str = ""


def extract_meta_from_filename(filepath: str) -> ElectionMeta:
    """Best-effort metadata from the filename and path."""
    fname = os.path.basename(filepath)
    parent = os.path.basename(os.path.dirname(filepath))
    grandparent = os.path.basename(os.path.dirname(os.path.dirname(filepath)))
    combined = grandparent + "/" + parent + "/" + fname
    meta = ElectionMeta(source_file=fname)

    # ROC year
    m = _YEAR_RE.search(combined)
    if m:
        meta.roc_year = int(m.group(1))
        meta.ad_year = meta.roc_year + 1911

    # AD year fallback
    if not meta.ad_year:
        m = _AD_YEAR_RE.search(combined)
        if m:
            meta.ad_year = int(m.group(0))
            meta.roc_year = meta.ad_year - 1911

    # County / scope
    m = _COUNTY_RE.search(combined)
    if m:
        meta.scope = m.group(1)

    # Region code
    m = _CODE_RE.search(fname)
    if m:
        meta.region_code = m.group(1)

    # Election type detection
    text = combined.lower()
    if "總統" in combined:
        meta.election_type = "president"
    elif "不分區" in combined or "政黨票" in combined or "全國不分區" in combined:
        meta.election_type = "legislator_party"
    elif "立法委員" in combined or "立委" in combined:
        meta.election_type = "legislator_regional"
    elif "直轄市長" in combined or ("市長" in combined and "縣市長" not in combined):
        meta.election_type = "mayor"
    elif "縣市長" in combined or "縣長" in combined:
        meta.election_type = "county_head"
    elif "鄉鎮市長" in combined:
        meta.election_type = "township_head"
    elif "議員" in combined:
        meta.election_type = "council"
    elif "鄉鎮市民代表" in combined or "代表" in combined:
        meta.election_type = "township_rep"
    elif "村里長" in combined:
        meta.election_type = "village_chief"
    elif "公投" in combined or "公民投票" in combined:
        meta.election_type = "referendum"

    # Build name
    roc_str = f"{meta.roc_year}年" if meta.roc_year else ""
    type_names = {
        "president": "總統副總統選舉",
        "legislator_regional": "區域立法委員選舉",
        "legislator_party": "不分區立法委員選舉",
        "mayor": "直轄市長選舉",
        "county_head": "縣市長選舉",
        "township_head": "鄉鎮市長選舉",
        "council": "縣市議員選舉",
        "township_rep": "鄉鎮市民代表選舉",
        "village_chief": "村里長選舉",
        "referendum": "公民投票",
    }
    scope_str = meta.scope or "全國"
    meta.name = f"{roc_str}{scope_str}{type_names.get(meta.election_type, '選舉')}"

    if not meta.scope and meta.election_type in ("president", "legislator_party", "referendum"):
        meta.scope = "全國"

    return meta


# ── Parsed output structure ─────────────────────────────────

@dataclass
class ParsedCandidate:
    name: str
    number: int = 0
    party: str = ""
    running_mate: str = ""  # for president


@dataclass
class ParsedRow:
    """One row of parsed vote data."""
    county: str = ""
    district: str = ""
    village: str = ""
    polling_station: str = ""
    candidate_name: str = ""
    candidate_number: int = 0
    party: str = ""
    vote_count: int = 0
    # Stats fields (only on first candidate per region)
    valid_votes: int = 0
    invalid_votes: int = 0
    votes_cast: int = 0
    eligible_voters: int = 0
    turnout: float | None = None
    ballots_issued: int = 0
    unreturned: int = 0
    remaining: int = 0


# ── Party name cleaning ─────────────────────────────────────

_PARTY_ALIASES = {
    "國": "中國國民黨", "國民黨": "中國國民黨",
    "民": "民主進步黨", "民進黨": "民主進步黨",
    "眾": "台灣民眾黨", "民眾黨": "台灣民眾黨", "臺灣民眾黨": "台灣民眾黨",
    "親": "親民黨",
    "時": "時代力量", "時力": "時代力量",
    "基": "台灣基進", "臺灣基進": "台灣基進",
    "新": "新黨",
    "台聯": "台灣團結聯盟", "臺灣團結聯盟": "台灣團結聯盟",
    "綠": "綠黨",
    "無": "無黨籍及未經政黨推薦",
    "無黨籍": "無黨籍及未經政黨推薦",
    "無黨": "無黨籍及未經政黨推薦",
}


def _normalize_party(raw: str) -> str:
    """Normalize party name to canonical form."""
    raw = raw.strip()
    if not raw:
        return "無黨籍及未經政黨推薦"
    return _PARTY_ALIASES.get(raw, raw)


# ── Parser 1: XLS polling-station detail ────────────────────
# Format: 縣表3-1-XXXXX(縣市名)-111年...xls
# Multi-header: rows 0-3 are title/headers, row 4+ data
# Columns: 行政區別 | 村里別 | 投開票所別 | candidate1 | candidate2 | ... | 有效票 | 無效票 | ...

def parse_xls(filepath: str, meta: ElectionMeta | None = None) -> tuple[ElectionMeta, list[ParsedRow]]:
    """Parse CEC .xls polling-station detail file."""
    import xlrd

    if meta is None:
        meta = extract_meta_from_filename(filepath)

    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)

    # Detect A05-2 format (district-level summary, not polling-station detail)
    # A05-2 has "各組候選人得票情形" in header row and NO 村里別/投開票所 columns
    first_rows_text = " ".join(str(ws.cell_value(r, c)) for r in range(min(3, ws.nrows)) for c in range(min(5, ws.ncols)))
    if "各組候選人" in first_rows_text or ("A05-2" in os.path.basename(filepath)):
        # Convert XLS to list-of-lists and delegate to parse_xlsx logic
        wb.release_resources()
        return _parse_xls_as_a05(filepath, meta)

    # Find the header rows — look for '行政區別' or '鄉(鎮' pattern
    header_row = -1
    for r in range(min(10, ws.nrows)):
        row_vals = [str(ws.cell_value(r, c)).strip() for c in range(min(5, ws.ncols))]
        if any("行政區" in v or "鄉(鎮" in v or "鄉（鎮" in v for v in row_vals):
            header_row = r
            break

    if header_row < 0:
        logger.warning(f"Cannot find header row in {filepath}")
        return meta, []

    # Parse candidate info from header rows (usually header_row + 1 or header_row + 2)
    # Candidates are in columns after the first 3 (district, village, polling station)
    # Format in header: "(1)\n候選人名\n政黨"
    candidates: list[ParsedCandidate] = []
    cand_cols: list[int] = []

    # Detect column layout
    header_vals = [str(ws.cell_value(header_row, c)).strip() for c in range(ws.ncols)]

    # Find where candidates start and stats columns begin
    stats_start_col = -1
    for c in range(3, ws.ncols):
        val = header_vals[c] if c < len(header_vals) else ""
        if any(k in val for k in ["有效票", "A\n有效", "A\r\n有效"]):
            stats_start_col = c
            break

    if stats_start_col < 0:
        # Try second approach: find "各候選人" header spanning candidates
        for c in range(3, ws.ncols):
            val = header_vals[c]
            if "有效" in val or "無效" in val:
                stats_start_col = c
                break

    if stats_start_col < 0:
        stats_start_col = ws.ncols  # no stats columns found

    # Parse candidate columns (between col 3 and stats_start_col)
    # Sub-header row contains candidate number, name, party
    sub_row = header_row + 1
    if sub_row >= ws.nrows:
        return meta, []

    for c in range(3, stats_start_col):
        cell_val = str(ws.cell_value(sub_row, c)).strip()
        if not cell_val:
            continue

        # Parse: "(1)\n柯文哲\n吳欣盈" or "1\n陳美妃\n無"
        lines = re.split(r"[\n\r]+", cell_val)
        lines = [_strip_ws(l) for l in lines if _strip_ws(l)]

        num = 0
        name = ""
        party = ""
        mate = ""

        for line in lines:
            # Number line: "(1)" or "1"
            m = re.match(r"\(?(\d+)\)?$", line)
            if m:
                num = int(m.group(1))
                continue
            # Name or party
            if not name:
                name = line
            elif not party and not mate:
                # Could be running mate or party
                # If the candidate has 2 names (president election), second is running mate
                if meta.election_type == "president" and len(line) >= 2 and not any(
                    k in line for k in ["黨", "無", "聯盟", "力量", "基進"]
                ):
                    mate = line
                else:
                    party = line
            elif not party:
                party = line

        if name:
            candidates.append(ParsedCandidate(
                name=name, number=num,
                party=_normalize_party(party),
                running_mate=mate,
            ))
            cand_cols.append(c)

    if not candidates:
        logger.warning(f"No candidates found in {filepath}")
        return meta, []

    # Map stats columns
    stats_map: dict[str, int] = {}
    for c in range(stats_start_col, ws.ncols):
        val = str(ws.cell_value(header_row, c)).strip()
        sub = str(ws.cell_value(sub_row, c)).strip() if sub_row < ws.nrows else ""
        combined = val + sub
        if "有效票" in combined and "無效" not in combined:
            stats_map["valid"] = c
        elif "無效票" in combined:
            stats_map["invalid"] = c
        elif "投票數" in combined:
            stats_map["cast"] = c
        elif "選舉人" in combined:
            stats_map["eligible"] = c
        elif "投票率" in combined:
            stats_map["turnout"] = c
        elif "發出票" in combined:
            stats_map["issued"] = c
        elif "已領未投" in combined:
            stats_map["unreturned"] = c
        elif "用餘票" in combined:
            stats_map["remaining"] = c

    # Parse data rows (skip header + sub-header + possible blank rows)
    rows: list[ParsedRow] = []
    current_district = ""
    current_village = ""
    data_start = sub_row + 1

    # Skip blank rows after sub-header
    while data_start < ws.nrows:
        first_cells = [str(ws.cell_value(data_start, c)).strip() for c in range(min(4, ws.ncols))]
        if any(first_cells):
            break
        data_start += 1

    for r in range(data_start, ws.nrows):
        col0 = _strip_ws(str(ws.cell_value(r, 0)))
        col1 = _strip_ws(str(ws.cell_value(r, 1)))
        col2 = _strip_ws(str(ws.cell_value(r, 2)))

        # Skip total/summary rows
        if any(k in col0 for k in ["總　計", "總計", "合　計", "合計"]):
            continue

        # Update district/village
        if col0:
            current_district = col0
            current_village = ""
        if col1:
            current_village = col1

        polling = col2 if col2 else ""

        # Skip if no candidate data
        if not current_district:
            continue

        # Check if any candidate column has data
        has_data = False
        for ci, c in enumerate(cand_cols):
            v = _clean_number(ws.cell_value(r, c))
            if v > 0:
                has_data = True
                break
        if not has_data:
            continue

        county = meta.scope or ""

        # Parse stats
        valid = _clean_number(ws.cell_value(r, stats_map["valid"])) if "valid" in stats_map else 0
        invalid = _clean_number(ws.cell_value(r, stats_map["invalid"])) if "invalid" in stats_map else 0
        cast = _clean_number(ws.cell_value(r, stats_map["cast"])) if "cast" in stats_map else 0
        eligible = _clean_number(ws.cell_value(r, stats_map["eligible"])) if "eligible" in stats_map else 0
        turnout = _clean_float(ws.cell_value(r, stats_map["turnout"])) if "turnout" in stats_map else None
        issued = _clean_number(ws.cell_value(r, stats_map["issued"])) if "issued" in stats_map else 0
        unreturned = _clean_number(ws.cell_value(r, stats_map["unreturned"])) if "unreturned" in stats_map else 0
        remaining = _clean_number(ws.cell_value(r, stats_map["remaining"])) if "remaining" in stats_map else 0

        for ci, cand in enumerate(candidates):
            c = cand_cols[ci]
            vote = _clean_number(ws.cell_value(r, c))
            rows.append(ParsedRow(
                county=county,
                district=current_district,
                village=current_village if polling else "",
                polling_station=polling,
                candidate_name=cand.name,
                candidate_number=cand.number,
                party=cand.party,
                vote_count=vote,
                valid_votes=valid if ci == 0 else 0,
                invalid_votes=invalid if ci == 0 else 0,
                votes_cast=cast if ci == 0 else 0,
                eligible_voters=eligible if ci == 0 else 0,
                turnout=turnout if ci == 0 else None,
                ballots_issued=issued if ci == 0 else 0,
                unreturned=unreturned if ci == 0 else 0,
                remaining=remaining if ci == 0 else 0,
            ))

    logger.info(f"Parsed {filepath}: {len(candidates)} candidates, {len(rows)} rows")
    return meta, rows


# ── Known presidential candidates → party mapping ──────────
# Used when XLSX format doesn't include party in the cell

_PRESIDENT_PARTY_MAP = {
    # 2024 (第16任)
    "柯文哲": "台灣民眾黨", "賴清德": "民主進步黨", "侯友宜": "中國國民黨",
    # 2020 (第15任)
    "蔡英文": "民主進步黨", "韓國瑜": "中國國民黨", "宋楚瑜": "親民黨",
    # 2016 (第14任)
    "朱立倫": "中國國民黨",
    # 2012 (第13任)
    "馬英九": "中國國民黨", "蔡英文": "民主進步黨", "宋楚瑜": "親民黨",
    # 2008 (第12任)
    "馬英九": "中國國民黨", "謝長廷": "民主進步黨",
    # 2004 (第11任)
    "陳水扁": "民主進步黨", "連戰": "中國國民黨",
    # 2000 (第10任)
    "陳水扁": "民主進步黨", "連戰": "中國國民黨", "宋楚瑜": "無黨籍及未經政黨推薦",
    "李敖": "新黨", "許信良": "無黨籍及未經政黨推薦",
    # 1996 (第9任)
    "李登輝": "中國國民黨", "彭明敏": "民主進步黨", "林洋港": "無黨籍及未經政黨推薦",
    "陳履安": "無黨籍及未經政黨推薦",
}


def _lookup_president_party(name: str) -> str:
    return _PRESIDENT_PARTY_MAP.get(name, "")


# ── XLS → A05 bridge: read XLS into list-of-lists, then use parse_xlsx logic ──

def _parse_xls_as_a05(filepath: str, meta: ElectionMeta | None = None) -> tuple[ElectionMeta, list[ParsedRow]]:
    """Read an XLS file that uses A05-2/A05-3 format and parse it via XLSX logic."""
    import xlrd
    if meta is None:
        meta = extract_meta_from_filename(filepath)

    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)

    # Convert to list-of-lists (same as openpyxl read)
    all_rows = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            val = ws.cell_value(r, c)
            # xlrd returns floats for numbers; keep as-is
            row.append(val if val != "" else None)
        all_rows.append(row)
    wb.release_resources()

    # Now run the XLSX parsing logic directly
    return _parse_a05_rows(all_rows, filepath, meta)


# ── Parser 2: XLSX summary (A05-2 format) ──────────────────
# Format: 總統-A05-2-候選人得票數一覽表(臺中市).xlsx
# Similar structure to XLS but different header format

def parse_xlsx(filepath: str, meta: ElectionMeta | None = None) -> tuple[ElectionMeta, list[ParsedRow]]:
    """Parse CEC .xlsx candidate vote summary."""
    import openpyxl

    if meta is None:
        meta = extract_meta_from_filename(filepath)

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active

    # Read all rows into memory for easier processing
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([v for v in row])
    wb.close()

    return _parse_a05_rows(all_rows, filepath, meta)


def _parse_a05_rows(all_rows: list[list], filepath: str, meta: ElectionMeta | None = None) -> tuple[ElectionMeta, list[ParsedRow]]:
    """Core parsing logic for A05-2/A05-3 format (shared by XLSX and XLS)."""
    if meta is None:
        meta = extract_meta_from_filename(filepath)

    if not all_rows:
        return meta, []

    # Extract county and year from sheet title or first row
    title_row = str(all_rows[0][0]) if all_rows[0][0] else ""
    m = _COUNTY_RE.search(title_row)
    if m and not meta.scope:
        meta.scope = m.group(1)

    # Try to get year from title: "第16任" → 2024, "第15任" → 2020
    if not meta.ad_year:
        m_term = re.search(r"第(\d+)任", title_row)
        if m_term:
            term = int(m_term.group(1))
            _pres_years = {9: 1996, 10: 2000, 11: 2004, 12: 2008, 13: 2012, 14: 2016, 15: 2020, 16: 2024}
            if term in _pres_years:
                meta.ad_year = _pres_years[term]
                meta.roc_year = meta.ad_year - 1911
        if not meta.ad_year:
            # Try "第N屆" for legislator terms
            m_leg = re.search(r"第(\d+)屆", title_row)
            if m_leg:
                leg_term = int(m_leg.group(1))
                _leg_years = {7: 2008, 8: 2012, 9: 2016, 10: 2020, 11: 2024}
                if leg_term in _leg_years:
                    meta.ad_year = _leg_years[leg_term]
                    meta.roc_year = meta.ad_year - 1911
        if not meta.ad_year:
            m_yr = _YEAR_RE.search(title_row)
            if m_yr:
                meta.roc_year = int(m_yr.group(1))
                meta.ad_year = meta.roc_year + 1911

    # Find header row — must have MULTIPLE non-empty cells (title rows have only 1)
    header_idx = -1
    for i, row in enumerate(all_rows[:10]):
        non_empty = sum(1 for v in row if v is not None and str(v).strip())
        row_text = " ".join(str(v) for v in row if v)
        if non_empty >= 3 and ("鄉" in row_text or "行政區" in row_text or "區別" in row_text):
            header_idx = i
            break

    if header_idx < 0:
        logger.warning(f"Cannot find header in {filepath}")
        return meta, []

    # Parse candidates from sub-header.
    # The candidate row may be header_idx+1 or header_idx+2 depending on format.
    # Detect which row actually has candidate info (contains "(1)" or number pattern).
    sub_idx = header_idx + 1
    if sub_idx >= len(all_rows):
        return meta, []

    # Check if sub_idx row has candidate data, else try sub_idx+1
    sub_text = " ".join(str(v) for v in all_rows[sub_idx] if v)
    if not re.search(r"\(\d+\)", sub_text):
        # Try next row
        if sub_idx + 1 < len(all_rows):
            alt_text = " ".join(str(v) for v in all_rows[sub_idx + 1] if v)
            if re.search(r"\(\d+\)", alt_text):
                sub_idx = sub_idx + 1

    header_vals = [str(v).strip() if v else "" for v in all_rows[header_idx]]
    sub_vals = [str(v).strip() if v else "" for v in all_rows[sub_idx]]

    # Find candidate columns and stats columns
    candidates: list[ParsedCandidate] = []
    cand_cols: list[int] = []
    stats_map: dict[str, int] = {}

    # Detect columns: use header row for stats, sub row for candidates
    ncols = max(len(header_vals), len(sub_vals))
    for c in range(ncols):
        hdr = header_vals[c] if c < len(header_vals) else ""
        sub = sub_vals[c] if c < len(sub_vals) else ""

        # Stats detection — primarily from header row
        # Order matters: check more-specific patterns first to avoid substring matches
        # e.g. "已領未投票數" contains "投票數", so check "已領未投" first
        if "已領未投" in hdr:
            stats_map["unreturned"] = c
            continue
        elif "有效票" in hdr and "無效" not in hdr:
            stats_map["valid"] = c
            continue
        elif "無效票" in hdr:
            stats_map["invalid"] = c
            continue
        elif "投票數" in hdr:
            stats_map["cast"] = c
            continue
        elif "投票率" in hdr:
            stats_map["turnout"] = c
            continue
        elif "選舉人" in hdr:
            stats_map["eligible"] = c
            continue
        elif "發出票" in hdr:
            stats_map["issued"] = c
            continue
        elif "用餘票" in hdr:
            stats_map["remaining"] = c
            continue

        # Candidate detection from sub row: "(1)\n柯文哲\n吳欣盈"
        lines = re.split(r"[\n\r]+", sub)
        lines = [_strip_ws(l) for l in lines if _strip_ws(l)]
        if not lines:
            continue

        num = 0
        name = ""
        party = ""
        mate = ""
        for line in lines:
            m_num = re.match(r"\(?(\d+)\)?$", line)
            if m_num:
                num = int(m_num.group(1))
                continue
            if not name:
                name = line
            elif meta.election_type == "president" and not mate and len(line) >= 2 and not any(
                k in line for k in ["黨", "無", "聯盟", "力量", "基進"]
            ):
                mate = line
            elif not party:
                party = line

        if name and c >= 1:  # skip col 0 (district name)
            # For presidential elections, party is often missing from the cell.
            # Use a known-candidate lookup.
            if not party and meta.election_type == "president":
                party = _lookup_president_party(name)
            candidates.append(ParsedCandidate(name=name, number=num, party=_normalize_party(party), running_mate=mate))
            cand_cols.append(c)

    if not candidates:
        logger.warning(f"No candidates found in {filepath}")
        return meta, []

    # Parse data rows
    rows: list[ParsedRow] = []
    data_start = sub_idx + 1

    # Skip blanks
    while data_start < len(all_rows):
        if any(all_rows[data_start]):
            break
        data_start += 1

    for r_idx in range(data_start, len(all_rows)):
        row_data = all_rows[r_idx]
        col0 = _strip_ws(str(row_data[0])) if row_data[0] else ""

        if any(k in col0 for k in ["總　計", "總計", "合　計", "合計"]):
            continue
        if not col0:
            continue

        county = meta.scope or ""
        district = col0

        valid = _clean_number(row_data[stats_map["valid"]]) if "valid" in stats_map and stats_map["valid"] < len(row_data) else 0
        invalid = _clean_number(row_data[stats_map["invalid"]]) if "invalid" in stats_map and stats_map["invalid"] < len(row_data) else 0
        cast = _clean_number(row_data[stats_map["cast"]]) if "cast" in stats_map and stats_map["cast"] < len(row_data) else 0
        eligible = _clean_number(row_data[stats_map["eligible"]]) if "eligible" in stats_map and stats_map["eligible"] < len(row_data) else 0
        turnout = _clean_float(row_data[stats_map["turnout"]]) if "turnout" in stats_map and stats_map["turnout"] < len(row_data) else None
        issued = _clean_number(row_data[stats_map["issued"]]) if "issued" in stats_map and stats_map["issued"] < len(row_data) else 0
        unreturned = _clean_number(row_data[stats_map["unreturned"]]) if "unreturned" in stats_map and stats_map["unreturned"] < len(row_data) else 0
        remaining = _clean_number(row_data[stats_map["remaining"]]) if "remaining" in stats_map and stats_map["remaining"] < len(row_data) else 0

        for ci, cand in enumerate(candidates):
            c = cand_cols[ci]
            vote = _clean_number(row_data[c]) if c < len(row_data) else 0
            rows.append(ParsedRow(
                county=county, district=district,
                candidate_name=cand.name, candidate_number=cand.number,
                party=cand.party, vote_count=vote,
                valid_votes=valid if ci == 0 else 0,
                invalid_votes=invalid if ci == 0 else 0,
                votes_cast=cast if ci == 0 else 0,
                eligible_voters=eligible if ci == 0 else 0,
                turnout=turnout if ci == 0 else None,
                ballots_issued=issued if ci == 0 else 0,
                unreturned=unreturned if ci == 0 else 0,
                remaining=remaining if ci == 0 else 0,
            ))

    logger.info(f"Parsed {filepath}: {len(candidates)} candidates, {len(rows)} rows")
    return meta, rows


# ── Parser 3: CSV/TXT aggregated ────────────────────────────
# Format: Big5 CSV or tab-separated, first row is header with candidate names in parentheses
# e.g. "行政區別,村里別,投開票所別,陳美妃（無）,蔡其昌（民主進步黨）,...,投票率"

_CAND_PAREN_RE = re.compile(r"^(.+?)[（(](.+?)[）)]$")


def parse_csv(filepath: str, meta: ElectionMeta | None = None) -> tuple[ElectionMeta, list[ParsedRow]]:
    """Parse CEC CSV/TXT aggregated file."""
    if meta is None:
        meta = extract_meta_from_filename(filepath)

    enc = _detect_encoding(filepath)

    with open(filepath, encoding=enc) as f:
        content = f.read()

    # Detect delimiter
    first_line = content.split("\n")[0]
    if "\t" in first_line:
        delimiter = "\t"
    else:
        delimiter = ","

    reader = csv.reader(io.StringIO(content), delimiter=delimiter)
    header = next(reader, None)
    if not header:
        return meta, []

    header = [_strip_ws(h) for h in header]

    # Find candidate columns (those with parenthesized party names)
    candidates: list[ParsedCandidate] = []
    cand_cols: list[int] = []
    turnout_col = -1
    district_col = -1
    village_col = -1
    ps_col = -1

    for i, h in enumerate(header):
        if h in ("行政區別", "行政區"):
            district_col = i
        elif h in ("村里別", "村里"):
            village_col = i
        elif h in ("投開票所別", "投開票所"):
            ps_col = i
        elif h in ("投票率",):
            turnout_col = i
        else:
            # Try candidate pattern: "候選人名（政黨）" or just "候選人名"
            m = _CAND_PAREN_RE.match(h)
            if m:
                name = _strip_ws(m.group(1))
                party = _normalize_party(m.group(2))
                candidates.append(ParsedCandidate(name=name, party=party, number=len(candidates) + 1))
                cand_cols.append(i)
            elif i > 0 and h and "票" not in h and "率" not in h and "選舉" not in h and "投票" not in h:
                # Bare candidate name (no party), e.g. "宋原通" in TXT
                candidates.append(ParsedCandidate(name=h, party="無黨籍及未經政黨推薦", number=len(candidates) + 1))
                cand_cols.append(i)

    if not candidates:
        logger.warning(f"No candidates found in CSV {filepath}")
        return meta, []

    rows: list[ParsedRow] = []
    county = meta.scope or ""

    for data_row in reader:
        if not data_row or len(data_row) < len(header):
            continue

        district = _strip_ws(data_row[district_col]) if district_col >= 0 else ""
        village = _strip_ws(data_row[village_col]) if village_col >= 0 else ""
        ps = _strip_ws(data_row[ps_col]) if ps_col >= 0 else ""
        turnout = _clean_float(data_row[turnout_col]) if turnout_col >= 0 else None

        if not district:
            continue
        if any(k in district for k in ["總計", "合計"]):
            continue

        for ci, cand in enumerate(candidates):
            c = cand_cols[ci]
            vote = _clean_number(data_row[c])
            rows.append(ParsedRow(
                county=county, district=district,
                village=village, polling_station=ps,
                candidate_name=cand.name, candidate_number=cand.number,
                party=cand.party, vote_count=vote,
                turnout=turnout if ci == 0 else None,
            ))

    logger.info(f"Parsed CSV {filepath}: {len(candidates)} candidates, {len(rows)} rows")
    return meta, rows


# ── Parser 4: 選舉結果清冊 (縣表11 / A06) ──────────────────
# Format: Each row = one candidate. Columns: 縣市別, 號次, 姓名, 性別, 出生年月日, 政黨, 得票數, 當選
# Multi-sheet: one sheet per county

def parse_result_registry(filepath: str, meta: ElectionMeta | None = None) -> tuple[ElectionMeta, list[ParsedRow]]:
    """Parse 選舉結果清冊 (縣表11 / A06)."""
    if meta is None:
        meta = extract_meta_from_filename(filepath)

    ext = Path(filepath).suffix.lower()
    sheets_data: list[tuple[str, list[list]]] = []

    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(filepath)
        for ws in wb.sheets():
            all_rows = []
            for r in range(ws.nrows):
                all_rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
            sheets_data.append((ws.name, all_rows))
    else:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        for ws in wb.worksheets:
            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append(list(row))
            sheets_data.append((ws.title, all_rows))
        wb.close()

    rows: list[ParsedRow] = []

    for sheet_name, sheet_rows in sheets_data:
        if not sheet_rows:
            continue

        # Find header row with '號次' or '姓名'
        header_idx = -1
        for i, row in enumerate(sheet_rows[:10]):
            row_text = " ".join(str(v) for v in row if v)
            if "號次" in row_text and ("姓名" in row_text or "候選人" in row_text):
                header_idx = i
                break
        if header_idx < 0:
            continue

        header = [str(v).strip() if v else "" for v in sheet_rows[header_idx]]

        # Map columns
        col_map: dict[str, int] = {}
        for i, h in enumerate(header):
            if "縣市" in h or "地區" in h or "行政區" in h or "選區" in h:
                col_map["region"] = i
            elif h == "號次":
                col_map["number"] = i
            elif "姓名" in h or "候選人" in h:
                col_map["name"] = i
            elif "政黨" in h or "推薦" in h:
                col_map["party"] = i
            elif "得票數" in h:
                col_map["votes"] = i
            elif "得票率" in h:
                col_map["share"] = i
            elif "當選" in h:
                col_map["elected"] = i
            elif "現任" in h:
                col_map["incumbent"] = i
            elif "性別" in h:
                col_map["gender"] = i

        if "name" not in col_map or "votes" not in col_map:
            continue

        current_region = sheet_name  # fallback: sheet name is often the county
        for r_idx in range(header_idx + 1, len(sheet_rows)):
            row = sheet_rows[r_idx]
            if not row or len(row) <= max(col_map.values()):
                continue

            name = _strip_ws(str(row[col_map["name"]])) if row[col_map["name"]] else ""
            if not name:
                continue

            # Region
            if "region" in col_map:
                rgn = _strip_ws(str(row[col_map["region"]])) if row[col_map["region"]] else ""
                if rgn:
                    current_region = rgn

            number = int(float(row[col_map["number"]])) if "number" in col_map and row[col_map["number"]] else 0
            party_raw = _strip_ws(str(row[col_map["party"]])) if "party" in col_map and row[col_map["party"]] else ""
            party = _normalize_party(party_raw)
            vote_count = _clean_number(row[col_map["votes"]])

            # Parse county from region (e.g. "臺中市" or "連江縣第01選區")
            county_match = _COUNTY_RE.search(current_region)
            county = county_match.group(1) if county_match else current_region

            rows.append(ParsedRow(
                county=county,
                district="",  # county-level totals
                candidate_name=name,
                candidate_number=number,
                party=party,
                vote_count=vote_count,
            ))

    if not meta.ad_year and rows:
        # Try from title
        title = str(sheets_data[0][1][0][0]) if sheets_data and sheets_data[0][1] else ""
        m = _YEAR_RE.search(title)
        if m:
            meta.roc_year = int(m.group(1))
            meta.ad_year = meta.roc_year + 1911

    logger.info(f"Parsed registry {filepath}: {len(rows)} candidate entries")
    return meta, rows


# ── Parser 5: 中表2 — 政黨得票率表 ─────────────────────────
# Format: Rows = counties, Columns = parties with 得票數/得票率 pairs

def parse_party_votes(filepath: str, meta: ElectionMeta | None = None) -> tuple[ElectionMeta, list[ParsedRow]]:
    """Parse 中表2 政黨得票率表."""
    if meta is None:
        meta = extract_meta_from_filename(filepath)

    import xlrd
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)

    # Row 0: title, Row 1: county header + party names, Row 2: 得票數/得票率 sub-header
    # Find header row
    header_idx = -1
    for r in range(min(5, ws.nrows)):
        row_text = " ".join(str(ws.cell_value(r, c)) for c in range(ws.ncols) if ws.cell_value(r, c))
        if "行政區" in row_text and ("國民黨" in row_text or "民進黨" in row_text or "政黨" in row_text):
            header_idx = r
            break

    if header_idx < 0:
        logger.warning(f"Cannot find header in party votes table {filepath}")
        return meta, []

    header = [str(ws.cell_value(header_idx, c)).strip() for c in range(ws.ncols)]

    # Extract party names from header (skip first 1-2 columns which are region/total)
    # Parties appear in pairs: 得票數, 得票率
    parties: list[tuple[str, int]] = []  # (party_name, vote_count_col)
    for c in range(1, len(header)):
        h = header[c].strip()
        if h and h != "候選人總得票數" and "得票" not in h:
            # This is a party name; next column is 得票數
            sub_r = header_idx + 1 if header_idx + 1 < ws.nrows else header_idx
            sub_val = str(ws.cell_value(sub_r, c)).strip() if sub_r < ws.nrows else ""
            if "得票數" in sub_val:
                parties.append((_normalize_party(h), c))
            elif c + 1 < ws.ncols:
                # Party name might span, 得票數 might be same col
                parties.append((_normalize_party(h), c))

    if not parties:
        logger.warning(f"No parties found in {filepath}")
        return meta, []

    # Parse data rows
    rows: list[ParsedRow] = []
    data_start = header_idx + 2  # skip sub-header

    for r in range(data_start, ws.nrows):
        region = _strip_ws(str(ws.cell_value(r, 0)))
        if not region or "總　計" in region or "合　計" in region:
            continue
        if region in ("臺灣省", "福建省"):
            continue  # skip province-level aggregates

        county_match = _COUNTY_RE.search(region)
        county = county_match.group(1) if county_match else region

        for party_name, vcol in parties:
            vote_count = _clean_number(ws.cell_value(r, vcol))
            if vote_count > 0:
                rows.append(ParsedRow(
                    county=county,
                    district="",  # county-level
                    candidate_name=party_name,  # party as "candidate" for party-vote tables
                    party=party_name,
                    vote_count=vote_count,
                ))

    logger.info(f"Parsed party votes {filepath}: {len(parties)} parties, {len(rows)} rows")
    return meta, rows


# ── District-level aggregation ──────────────────────────────

def aggregate_to_district(rows: list[ParsedRow]) -> list[ParsedRow]:
    """Collapse village / polling-station rows into district-level summaries.

    Groups by (county, district, candidate_name) and sums vote_count
    and stats fields.  Turnout is recalculated from votes_cast / eligible_voters.
    """
    if not rows:
        return rows

    # Check if data is already at district level (no village/polling_station)
    has_detail = any(r.village or r.polling_station for r in rows)
    if not has_detail:
        return rows  # already district level, nothing to do

    from collections import defaultdict

    # Key: (county, district, candidate_name)
    vote_agg: dict[tuple, int] = defaultdict(int)
    # Key: (county, district)  — stats are per-region, not per-candidate
    stats_agg: dict[tuple, dict] = {}
    # Preserve candidate metadata
    cand_meta: dict[str, tuple] = {}  # name → (number, party)

    for r in rows:
        vk = (r.county, r.district, r.candidate_name)
        vote_agg[vk] += r.vote_count

        if r.candidate_name not in cand_meta:
            cand_meta[r.candidate_name] = (r.candidate_number, r.party)

        sk = (r.county, r.district)
        if sk not in stats_agg:
            stats_agg[sk] = {
                "valid_votes": 0, "invalid_votes": 0, "votes_cast": 0,
                "eligible_voters": 0, "ballots_issued": 0,
                "unreturned": 0, "remaining": 0,
            }
        s = stats_agg[sk]
        s["valid_votes"] += r.valid_votes
        s["invalid_votes"] += r.invalid_votes
        s["votes_cast"] += r.votes_cast
        s["eligible_voters"] += r.eligible_voters
        s["ballots_issued"] += r.ballots_issued
        s["unreturned"] += r.unreturned
        s["remaining"] += r.remaining

    # Build aggregated rows
    result: list[ParsedRow] = []
    # Sort by county, district, candidate number for stable output
    seen_districts: set[tuple] = set()

    for (county, district, cand_name), total_votes in sorted(vote_agg.items()):
        num, party = cand_meta[cand_name]
        sk = (county, district)
        s = stats_agg.get(sk, {})

        # Only attach stats to the first candidate per district
        is_first = sk not in seen_districts
        seen_districts.add(sk)

        eligible = s.get("eligible_voters", 0) if is_first else 0
        cast = s.get("votes_cast", 0) if is_first else 0
        turnout = round(cast / eligible * 100, 2) if is_first and eligible > 0 else None

        result.append(ParsedRow(
            county=county,
            district=district,
            village="",          # cleared — aggregated to district
            polling_station="",  # cleared
            candidate_name=cand_name,
            candidate_number=num,
            party=party,
            vote_count=total_votes,
            valid_votes=s.get("valid_votes", 0) if is_first else 0,
            invalid_votes=s.get("invalid_votes", 0) if is_first else 0,
            votes_cast=cast,
            eligible_voters=eligible,
            turnout=turnout,
            ballots_issued=s.get("ballots_issued", 0) if is_first else 0,
            unreturned=s.get("unreturned", 0) if is_first else 0,
            remaining=s.get("remaining", 0) if is_first else 0,
        ))

    logger.info(f"Aggregated {len(rows)} detail rows → {len(result)} district-level rows")
    return result


# ── Dispatcher ──────────────────────────────────────────────

def _detect_file_subtype(filepath: str) -> str:
    """Detect the sub-type of CEC data file based on filename patterns.

    Returns: 'registry' | 'party_votes' | 'polling_detail' | 'summary' | 'skip'
    """
    fname = os.path.basename(filepath)

    # Files to skip
    if any(k in fname for k in ["完成時間", "完成登錄", "中表5", "中表6", "中表8"]):
        return "skip"

    # 選舉結果清冊 (縣表11, A06)
    if "縣表11" in fname or "結果清冊" in fname or "A06" in fname:
        return "registry"

    # 政黨得票率表 (中表2)
    if "中表2" in fname:
        return "party_votes"

    # 投開票所明細 (縣表3)
    if "縣表3" in fname or "投開票所" in fname or "A05-4" in fname:
        return "polling_detail"

    # A05-2 summary or CSV/TXT
    return "summary"


def parse_file(filepath: str, meta: ElectionMeta | None = None) -> tuple[ElectionMeta, list[ParsedRow]]:
    """Auto-detect format and parse a CEC election file.

    Always aggregates to district level (鄉鎮市區).
    """
    ext = Path(filepath).suffix.lower()
    subtype = _detect_file_subtype(filepath)

    if subtype == "skip":
        logger.info(f"Skipping non-vote file: {os.path.basename(filepath)}")
        return meta or ElectionMeta(), []

    if subtype == "registry":
        meta, rows = parse_result_registry(filepath, meta)
        return meta, rows  # already county-level, no aggregation needed

    if subtype == "party_votes":
        meta, rows = parse_party_votes(filepath, meta)
        return meta, rows  # already county-level

    # Standard formats: polling detail, summary tables, CSV
    if ext == ".xls":
        meta, rows = parse_xls(filepath, meta)
    elif ext == ".xlsx":
        meta, rows = parse_xlsx(filepath, meta)
    elif ext in (".csv", ".txt", ".tsv"):
        meta, rows = parse_csv(filepath, meta)
    else:
        logger.warning(f"Unsupported file format: {ext} ({filepath})")
        return meta or ElectionMeta(), []

    # Aggregate to district level — drop village/polling_station detail
    rows = aggregate_to_district(rows)
    return meta, rows
