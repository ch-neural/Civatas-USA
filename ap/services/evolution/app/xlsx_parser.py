"""
XLSX Parser — 解析 XLSX 統計檔案

Supports:
  - Individual-level records → auto-aggregate to district level
  - Codebook files → extract column label mappings
  - Multi-file upload → codebook + data auto-detection
"""
from __future__ import annotations

import io
import logging
import os
from collections import defaultdict
from typing import Any

log = logging.getLogger(__name__)


def _is_codebook(headers: list[str], sample_rows: list[list]) -> bool:
    """Heuristic: a codebook has metadata-like headers (標籤/值/計數) and
    few data rows with mixed types, not numeric-heavy."""
    h_set = set(str(h).strip() for h in headers if h)
    codebook_keywords = {"值", "計數", "百分比", "標籤", "位置", "類型",
                         "格式", "測量", "有效", "遺漏值", "標準屬性",
                         "集中趨勢和離差", "有效值", "遺漏"}
    # If headers or first-column values contain codebook keywords
    first_col = set()
    for row in sample_rows[:50]:
        if row and row[0]:
            first_col.add(str(row[0]).strip())
    overlap = codebook_keywords & (h_set | first_col)
    return len(overlap) >= 3


def parse_codebook(raw_bytes: bytes) -> dict[str, dict]:
    """Parse a codebook XLSX into a column → {code: label} mapping.

    Returns: {"column_name": {"label": "...", "values": {code: label, ...}}}
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    codebook: dict[str, dict] = {}
    current_var = None

    for row in rows:
        col0 = str(row[0]).strip() if row[0] is not None else ""
        col1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        col2 = row[2] if len(row) > 2 else None

        # New variable definition: non-empty first column, not a metadata keyword
        skip_keywords = {"", "標準屬性", "N", "集中趨勢和離差", "標籤值",
                         "有效值", "遺漏值", "編碼簿", "None"}
        if col0 and col0 not in skip_keywords:
            current_var = col0
            if current_var not in codebook:
                codebook[current_var] = {"label": "", "values": {}}

        if not current_var:
            continue

        # Column label
        if col1 == "標籤" and col2:
            codebook[current_var]["label"] = str(col2)

        # Value labels: rows where col0 is "有效值" or "標籤值", or
        # continuation rows with (None, code, label)
        if col0 in ("有效值", "標籤值"):
            if col1 and col2 is not None:
                codebook[current_var]["values"][str(col1)] = str(col2)
        elif col0 == "" and col1 and col2 is not None:
            # Continuation of value labels
            entry = codebook.get(current_var)
            if entry and entry["values"]:
                codebook[current_var]["values"][str(col1)] = str(col2)

    return codebook


def parse_data_xlsx(raw_bytes: bytes,
                    codebook: dict[str, dict] | None = None
                    ) -> dict[str, dict[str, Any]]:
    """Parse a data XLSX and aggregate to district/admin level.

    - Detects district columns (city, town, vill or similar)
    - Detects numeric columns for aggregation
    - Returns {admin_key: {field: aggregated_value}}
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(rows_iter)]

    # Identify district columns
    district_cols = _find_district_columns(headers, codebook)
    # Identify numeric columns for aggregation
    agg_cols = _find_aggregatable_columns(headers, codebook)

    log.info(f"District columns: {district_cols}")
    log.info(f"Aggregation columns: {agg_cols}")

    # Aggregate
    stats: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))
    row_count = 0

    for row in rows_iter:
        if not row or all(v is None for v in row):
            continue
        row_count += 1

        # Build admin key
        key_parts = []
        for col_idx in district_cols:
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                val_str = str(val).strip()
                # Apply codebook label if available
                col_name = headers[col_idx]
                if codebook and col_name in codebook:
                    mapped = codebook[col_name]["values"].get(val_str, val_str)
                    key_parts.append(mapped)
                else:
                    key_parts.append(val_str)
        admin_key = "|".join(key_parts) if key_parts else "unknown"

        # Collect numeric values
        for col_idx, col_name in agg_cols:
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                try:
                    stats[admin_key][col_name].append(float(val))
                except (ValueError, TypeError):
                    pass

    wb.close()

    # Compute aggregated stats
    result: dict[str, dict[str, Any]] = {}
    for admin_key, fields in stats.items():
        agg: dict[str, Any] = {"sample_size": 0}
        for field_name, values in fields.items():
            if not values:
                continue
            n = len(values)
            agg["sample_size"] = max(agg.get("sample_size", 0), n)

            # For binary fields (0/1), compute rate
            unique = set(values)
            if unique <= {0.0, 1.0}:
                rate = sum(values) / n if n > 0 else 0
                label = _get_field_label(field_name, codebook)
                agg[f"{label}_率"] = round(rate, 4)
            else:
                # Continuous: compute mean
                label = _get_field_label(field_name, codebook)
                agg[f"{label}_平均"] = round(sum(values) / n, 2)

        result[admin_key] = agg

    log.info(f"Aggregated {row_count} rows into {len(result)} districts")
    return result


def _find_district_columns(headers: list[str],
                            codebook: dict | None) -> list[int]:
    """Find column indices that represent administrative divisions.
    
    Prefers text-based columns (city, town, vill) over numeric ID columns
    (cityid, townid, villid) when both exist.
    """
    district_keywords = [
        "city", "town", "vill", "district",
        "縣市", "鄉鎮", "村里", "行政區",
        "縣市別", "鄉鎮市區",
    ]
    id_suffixes = ["id"]
    
    text_cols = []
    id_cols = []
    
    for i, h in enumerate(headers):
        h_lower = h.lower()
        matched = any(kw in h_lower for kw in district_keywords)
        if not matched and codebook and h in codebook:
            label = codebook[h].get("label", "")
            matched = any(kw in label for kw in district_keywords)
        if matched:
            # Separate text vs ID columns
            if any(h_lower.endswith(s) for s in id_suffixes) and not any(
                    ch in h for ch in "市縣鄉鎮區里"):
                id_cols.append(i)
            else:
                text_cols.append(i)

    # Prefer text columns; fall back to ID columns
    found = text_cols if text_cols else id_cols

    # If no explicit district columns found, check codebook for geo values
    if not found and codebook:
        for i, h in enumerate(headers):
            if h in codebook:
                vals = codebook[h].get("values", {})
                sample_vals = list(vals.values())[:5]
                geo_keywords = ["市", "縣", "區", "鄉", "鎮"]
                if any(any(gk in sv for gk in geo_keywords) for sv in sample_vals):
                    found.append(i)
                    break

    return found


def _find_aggregatable_columns(headers: list[str],
                                codebook: dict | None) -> list[tuple[int, str]]:
    """Find columns suitable for aggregation (numeric, meaningful)."""
    skip_keywords = {"id", "newid", "pollid", "neighb", "wec", "wen",
                     "egroup", "cegroup", "note", "year", "month", "date",
                     "sex", "sex_v"}
    # District columns are handled separately
    district_keywords = {"city", "town", "vill", "cityid", "townid", "villid",
                         "district", "縣市", "鄉鎮", "村里", "行政區"}

    result = []
    for i, h in enumerate(headers):
        h_lower = h.lower().replace("\ufeff", "")
        if h_lower in skip_keywords or h_lower in district_keywords:
            continue
        if not h or h_lower.endswith("id"):
            continue
        result.append((i, h))

    return result


def _get_field_label(field_name: str, codebook: dict | None) -> str:
    """Get a human-readable label for a field."""
    if codebook and field_name in codebook:
        label = codebook[field_name].get("label", "")
        if label:
            return label
    return field_name


# ── Multi-file handler ───────────────────────────────────────────────

def process_xlsx_upload(files: list[tuple[str, bytes]]
                         ) -> dict[str, dict[str, Any]]:
    """Process one or more XLSX files.

    Auto-detects codebook vs data file. If a codebook is found,
    uses it to decode the data file's column values.

    Args:
        files: list of (filename, raw_bytes) tuples

    Returns:
        Aggregated district-level statistics dict
    """
    import openpyxl

    codebook: dict[str, dict] | None = None
    data_bytes: bytes | None = None
    data_filename = ""

    for filename, raw in files:
        # Quick peek to detect codebook
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb[wb.sheetnames[0]]
        peek_rows = []
        headers_raw = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers_raw = list(row)
            peek_rows.append(list(row))
            if i >= 60:
                break
        wb.close()

        if headers_raw and _is_codebook(
                [str(h) for h in headers_raw if h], peek_rows[1:]):
            log.info(f"Detected codebook: {filename}")
            codebook = parse_codebook(raw)
        else:
            log.info(f"Detected data file: {filename}")
            data_bytes = raw
            data_filename = filename

    if data_bytes is None:
        raise ValueError("未偵測到資料檔案，請確認上傳的檔案包含數值資料")

    return parse_data_xlsx(data_bytes, codebook)


# ── Generic CSV parser ───────────────────────────────────────────────

def parse_generic_csv(raw: bytes) -> dict[str, dict[str, Any]]:
    """Parse a CSV file (any encoding) and auto-aggregate to district level.

    Reuses the same column detection and aggregation logic as XLSX parser.
    Returns {admin_key: {field: aggregated_value}}.
    """
    import csv

    # Decode with fallback
    text = None
    for enc in ("utf-8-sig", "big5", "cp950", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        raise ValueError("無法偵測 CSV 檔案編碼")

    reader = csv.reader(io.StringIO(text))
    headers = [h.strip().replace("\ufeff", "") for h in next(reader)]

    # Reuse XLSX detection logic
    district_cols = _find_district_columns(headers, None)
    agg_cols = _find_aggregatable_columns(headers, None)

    log.info(f"CSV district columns: {[headers[i] for i in district_cols]}")
    log.info(f"CSV aggregation columns: {[h for _, h in agg_cols]}")

    if not district_cols:
        raise ValueError("未偵測到行政區欄位（city/town/vill/縣市/鄉鎮/村里）")

    # First pass: collect all rows to determine which columns are actually numeric
    all_rows = list(reader)
    log.info(f"CSV total rows: {len(all_rows)}")

    # Probe which agg_cols are actually numeric (sample first 200 rows)
    numeric_cols = []
    for col_idx, col_name in agg_cols:
        numeric_count = 0
        sample_vals = set()
        for row in all_rows[:200]:
            if col_idx < len(row):
                val = row[col_idx].strip()
                if val:
                    try:
                        sample_vals.add(float(val))
                        numeric_count += 1
                    except ValueError:
                        pass
        # Consider numeric if >50% of non-empty samples parse as numbers
        if numeric_count >= 20:
            numeric_cols.append((col_idx, col_name))

    log.info(f"CSV numeric columns: {[h for _, h in numeric_cols]}")

    # Aggregate
    stats: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))
    row_count = 0

    for row in all_rows:
        if not row or not any(cell.strip() for cell in row):
            continue
        row_count += 1

        # Build admin key
        key_parts = []
        for col_idx in district_cols:
            val = row[col_idx].strip() if col_idx < len(row) else ""
            if val:
                key_parts.append(val)
        admin_key = "|".join(key_parts) if key_parts else "unknown"

        # Collect numeric values
        for col_idx, col_name in numeric_cols:
            if col_idx < len(row):
                val = row[col_idx].strip()
                if val:
                    try:
                        stats[admin_key][col_name].append(float(val))
                    except ValueError:
                        pass

    # Compute aggregated stats (same logic as XLSX)
    result: dict[str, dict[str, Any]] = {}
    for admin_key, fields in stats.items():
        agg: dict[str, Any] = {"sample_size": 0}
        for field_name, values in fields.items():
            if not values:
                continue
            n = len(values)
            agg["sample_size"] = max(agg.get("sample_size", 0), n)

            unique = set(values)
            if unique <= {0.0, 1.0}:
                # Binary field → rate
                rate = sum(values) / n if n > 0 else 0
                agg[f"{field_name}_率"] = round(rate, 4)
            else:
                # Continuous → mean
                agg[f"{field_name}_平均"] = round(sum(values) / n, 2)

        result[admin_key] = agg

    log.info(f"CSV aggregated {row_count} rows into {len(result)} districts")
    return result
